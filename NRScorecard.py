# !/usr/bin/python3
import os
import importlib
import subprocess
import sys

# colour coding system for output
blue = "\u001b[36;1m"
green = "\u001b[32;1m"
red = "\u001b[31;1m"
yellow = "\u001b[33;1m"
endc = "\033[0m"

os.system("clear")

def check_dependencies():
    # List of required dependencies
    dependencies = [
        'os',
        'importlib',
        'subprocess',
        'sys',
        'configparser',
        'readline',
        're',
        'json',
        'datetime',
        'calendar',
        'copy',
        'urllib',
        'requests',
        'tqdm',
        'fuzzywuzzy',
        'openpyxl',
        'pandas',
        'io',
        'dateutil.relativedelta'
    ]

    # Initialize an empty list for missing dependencies
    missing_dependencies = []

    # Check if each dependency is installed, and add it to the missing_dependencies list if not
    for dependency in dependencies:
        try:
            importlib.import_module(dependency)
        except ImportError:
            missing_dependencies.append(dependency)

    # If there are no missing dependencies, print a success message and return
    if not missing_dependencies:
        print(f"{green}All required dependencies are installed.{endc}")
        return

    # Check if pip is installed, and exit if it is not
    try:
        import pip
    except ImportError:
        print(f"{red}pip is not installed. Please install pip and try again.{endc}")
        sys.exit(1)

    # Print the pip version
    print(f"pip version: {pip.__version__}")

    # Print each missing dependency
    for dependency in missing_dependencies:
        print(f"{yellow}Dependency {dependency} is missing.{endc}")

    # Ask the user if they want to install the missing dependencies
    install_dependencies = input("Do you want to install the missing dependencies? (Y/n): ")

    # If the user chooses to install the missing dependencies, attempt to install them
    if install_dependencies.lower() == 'y' or install_dependencies == '':
        for dependency in missing_dependencies:
            print(f"{yellow}Attempting to install {dependency}...{endc}")
            subprocess.check_call([sys.executable, "-m", "pip", "install", dependency])
            print(f"{green}All required dependencies are installed.{endc}")
    # If the user chooses not to install the missing dependencies, print a message and exit
    else:
        print(f"{red}Missing dependencies were not installed. Please install them manually and try again.{endc}")
        sys.exit(1)

# Check and install dependencies before importing them
check_dependencies()

import configparser
import readline
import re
import json
from datetime import datetime, timedelta, timezone
from dateutil.relativedelta import relativedelta
import calendar
import pandas as pd
from io import StringIO
from copy import copy
from urllib.parse import quote_plus
import requests
from tqdm import tqdm
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


BASE_URL = 'https://jira.corp.adobe.com/plugins/servlet/eazybi/accounts/295/export/report'
REPORT_IDS = {'req': 11252, 'prov': 11253, 'csopm': 11254}

# Create Config File Parser to retreive API username/password/keys
config = configparser.ConfigParser()

config.read("config.ini")
user_name = config.get("accounts", "username")
password = config.get("secrets", "password")
api_key = config.get("keys", "api_key")

# API Variables and NRQL Timeout
url = "https://api.newrelic.com/graphql"
headers = {"Content-Type": "application/json", "API-Key": api_key}
timeout = 60

# Text Styling variables
under_line = "\033[4m"
normal_txt = "\033[0m"
bold_txt = "\033[1m"

# Lists for each service card metric to store the results and write to the excel sheet.
transactional_success_rate_results_per_channel = []
transactional_latency_results = []
batch_success_rate_results = [] 
transactional_success_rate_results_per_channel = []
transactional_latency_results = []
db_storage_space_results = []
db_bloat_results = []
sftp_storage_space_results = []
active_profiles_results = []
peak_volume_batch_results = []
peak_volume_transactional_throughput_per_hour_results = []
cpgn_ticket_count = []
cpgn_ticket_avg_res = []
cso_ticket_count = []
rca_res_count = []

# Placement of helper functions first in the code
def format_storage_values(value):
    if value >= 1_000_000:  # Value is in MB
        return f"{value / 1_000_000:,.1f}TB"
    elif value >= 1_000:  # Value is in GB
        return f"{value / 1_000:,.1f}TB"
    else:  # Value is in MB
        return f"{value:.2f}GB"
    
def round_up_nr_values_to_nearest_integer(value):
    if value >= 1_000_000:
        return f"{round(value / 1_000_000)}M"
    elif value >= 1_000:
        return f"{round(value / 1_000)}K"
    else:
        return f"{value:.0f}"
    
def format_nr_values(value):
    if value >= 1_000_000:
        return f"{value / 1_000_000:,.2f}M"
    elif value >= 1_000:
        return f"{value / 1_000:,.2f}K"
    else:
        return f"{value:.0f}"
    
def shorten_hostname(host):
    """This function shortens the host to just the name and instance number e.g. brp-mkt-prod8"""
    match = re.search(r'([a-zA-Z-]+)\d+', host)
    if match:
        host = match.group(0)
    return host

def capitalise_customer_name(name):
    return name.upper()

def get_user_input(prompt):
    # Use readline to read user input
    line = input(prompt)
    # Add the line to readline history
    readline.add_history(line)
    # Return the user input
    return line

def match_casing(input_str, string_list):
    """This function matches the casing of the input string with the casing of a string from a list"""
    # Convert the input string to lowercase
    input_str_lower = input_str.lower()

    # Iterate through the string list
    for item in string_list:
        # Compare the lowercase versions of the input string and the current item
        if item.lower() == input_str_lower:
            # Initialize an empty matched string
            matched_str = ""
            # Iterate through the characters in the current item
            for i in range(len(item)):
                # Check if the current character is uppercase
                if item[i].isupper():
                    # Add the corresponding character from the input string in uppercase
                    matched_str += input_str[i].upper()
                else:
                    # Add the corresponding character from the input string in lowercase
                    matched_str += input_str[i].lower()
            # Print the matched string with correct casing
            print(f"Matched string with correct casing: {matched_str}")
            return matched_str

    # Return None if no match is found
    return None

def find_best_match(user_input, source_list):
    """This function finds the best matching string from a list of strings based on fuzzy string matching"""
    # Define a list of fuzzy string matching methods from the fuzzywuzzy library
    fuzz_methods = [fuzz.ratio, fuzz.partial_ratio, fuzz.token_sort_ratio, fuzz.token_set_ratio]
    # Initialize the best matching score and the best matching string
    best_score = 0
    best_match = None

    # Iterate through the strings in the source list
    for name in source_list:
        # Initialize the current best matching score
        current_best_score = 0
        # Iterate through the fuzzy string matching methods
        for method in fuzz_methods:
            # Calculate the matching score using the current method
            score = method(user_input, name)
            # Update the current best matching score if the calculated score is higher
            if score > current_best_score:
                current_best_score = score
        # Update the overall best matching score and string if the current best score is higher
        if current_best_score > best_score:
            best_score = current_best_score
            best_match = name
        # Break the loop if the best matching score is 99 or higher
        if best_score >= 99:
            break

    # Return the best matching string and its matching score
    return best_match, best_score

def sort_results_by_hostname_number(hostname):
    # Extract numbers from the hostname
    numbers = re.findall(r'\d+', hostname)
    return int(numbers[0]) if numbers else -1

def retrieve_customer_dashboard_to_copy():
    """This function will pull the dashboard for JP Morgan Chase and we'll use this as our template to create further customer dashboards"""
    print("Retrieveing template dashboard")
    query = {
        "query": '{\n  actor {\n    entity(guid: "MTIwOTMyN3xWSVp8REFTSEJPQVJEfGRhOjIwMTc0NjU") {\n      ... on DashboardEntity {\n        name\n        permissions\n        pages {\n          name\n          widgets {\n            visualization {\n              id\n            }\n            title\n            layout {\n              row\n              width\n              height\n              column\n            }\n            rawConfiguration\n          }\n        }\n      }\n    }\n  }\n}\n',
        "variables": "",
    }

    requests_dashboard = requests.post(
        url=url, headers=headers, json=query, timeout=timeout
    )

    if requests_dashboard.status_code == 200:
        print(requests_dashboard.status_code, "- Success")
        jp_morgan_chase_json = requests_dashboard.json()
    else:
        raise Exception(
            f"Unable to retrieve dashboard template, Nerdgraph query failed with a {requests_dashboard.status_code}."
        )        
    return jp_morgan_chase_json

def take_new_customer_details(
    jp_morgan_chase_dashboard_template, customer_name, tenant_id):
    print("updating the customer name and tenant id values")
    
    return (
        jp_morgan_chase_dashboard_template.replace("jpjupiter-mkt-prod7-1", tenant_id+'-mkt-prod%')
        .replace("jpjupiter", tenant_id)
        .replace("JP Morgan Chase", customer_name)
        .replace("JP Morgan", customer_name)
    )

def create_dashboard_api_query(query, customer_name):
    requests_dashboard_creation = requests.post(url, headers=headers, data=json.dumps(query), timeout=timeout)
    if requests_dashboard_creation.status_code == 200:
        print(requests_dashboard_creation.status_code, "- Success")
        json_response = json.dumps(requests_dashboard_creation.json(), indent=2)
        print(json_response)
        grab_re_guid = re.search(r"M[A-Za-z0-9]+", json_response)
        guid_id = grab_re_guid.group()
        print(f"Dashboard for {customer_name} has been successfully created.")
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to create the dashboard, Nerdgraph query failed with a {requests_dashboard_creation.status_code}."
        )
    return guid_id


def dash_url(id,name):
    name_with_spaces = name.replace(" ", "%20").lower()
    url = f"https://one.newrelic.com/dashboards/detail/{id}?account=1209327&filters=%28name%20LIKE%20%27{name_with_spaces}%27%20OR%20id%20%3D%20%27{name_with_spaces}%27%20OR%20domainId%20%3D%20%27{name_with_spaces}%27%29&state=edca9851-d3e7-a231-cae0-2e8233adc72d"
    return url

def check_for_existing_dashboard(customer_name, new_onboard):
    query_template = f"""
{{
  actor {{
    entitySearch(query: "{{query_operator}}", options: {{caseSensitiveTagMatching: false}}) {{
      results {{
        entities {{
          guid
          entityType
          name
          ... on DashboardEntityOutline {{
            guid
            name
            createdAt
          }}
        }}
      }}
    }}
  }}
}}
"""
    queries = [
        query_template.replace("{query_operator}", f"name = '{customer_name} - ACMS Dashboard'"),
        query_template.replace("{query_operator}", f"name like '{customer_name} - ACMS'")
    ]

    # Initialize a variable for results
    results = None

    # Iterate through each query in the queries list
    for query in queries:
        # Make a POST request to the specified URL with the query as the payload
        response = requests.post(url, headers=headers, data=json.dumps({"query": query}), timeout=timeout)
        
        # Check if the response status code is 200 (success)
        if response.status_code == 200:
            # Extract the data from the JSON response
            data = response.json()
            results = data["data"]["actor"]["entitySearch"]["results"]["entities"]

            # stop when a result is found
            if results:
                break
        else:
            # Raise an exception if the query fails
            raise Exception(f"Unable to check for an existing dashboard, Nerdgraph query failed with a {response.status_code}.")

            
    # Print a message if no results are found
    if not results:
        print(f"{red}No dashboard was found for {blue}{customer_name}.{endc}")
        return None

    # Sort the results by the "createdAt" field in descending order
    sorted_results = sorted(results, key=lambda x: x["createdAt"], reverse=True)

    # Get the current time
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Iterate through the sorted results
    printed_statement = False
    for result in sorted_results:
        # Extract relevant data from the result
        created = result["createdAt"].replace("T", " ")[:19]
        guid_id = result["guid"]
        name = result["name"]

        # Compare the customer name in a case-insensitive manner
        condition_result = customer_name.lower() in name.lower()

        if not condition_result:
            if not printed_statement:
                print(f"{red}No dashboard was found for {blue}{customer_name}.{endc}")
                printed_statement = True

        # Check if the condition_result is True
        if condition_result:
            # Handle cases when new_onboard is "y"
            if new_onboard == "y" and created < current_time:
                print("duplicates found, removing older dashboard entries")
                print(f"{name}\n{created}\n{guid_id}\n")
                remove_existing_dashboards(guid_id)
                return guid_id
            
            # Handle cases when new_onboard is "n"
            elif new_onboard == "n":
                if guid_id:
                    print(f"{green}Dashboard exists for {blue}{customer_name}!{endc}")
                return guid_id

def remove_existing_dashboards(id):
    remove_dashboard_query = f"""
mutation {{
  dashboardDelete(guid: "{id}") {{
    errors {{
      description
      type
    }}
    status
  }}
}}
"""
    # Send the API request
    requests_remove_dashboard = requests.post(url, headers=headers, data=json.dumps({"query": remove_dashboard_query}),timeout=timeout)
    if requests_remove_dashboard.status_code == 200:
        json_response = json.dumps(requests_remove_dashboard.json(), indent=2)
        print(json_response)
        print("Duplicates have been removed")
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to remove existing dashboard, Nerdgraph query failed with a {requests_remove_dashboard.status_code}."
        ) 

def enter_customer_details():
    try:
        customer_name = get_user_input("Enter Customer Name (e.g. Aldi Sued) as it's defined in JIRA: ").rstrip()
        customer_name = capitalise_customer_name(customer_name)
        tenant_id = get_user_input("Enter the tenant id. For example, aldisued: ").lower().rstrip()
        customer_excel_filename = customer_name.lower().replace(" ","_").replace("-","_")
        return customer_name, tenant_id, customer_excel_filename
    except KeyboardInterrupt:
        sys.exit()


def create_new_dashboard(customer_name, tenant_id):
    retrieve_dashboard = retrieve_customer_dashboard_to_copy()

    updated_dashboard_json = take_new_customer_details(
        json.dumps(retrieve_dashboard), customer_name, tenant_id
    )
    jsonvariables = json.loads(updated_dashboard_json)["data"]["actor"]
    jsonvariables["dashboard"] = jsonvariables.pop("entity")

    create_dashboard_query = {
        "query": "mutation create($dashboard: DashboardInput!) {\n  dashboardCreate(accountId: 1209327, dashboard: $dashboard) {\n    entityResult {\n      guid\n      name\n    }\n    errors {\n      description\n    }\n  }\n}\n",
        "variables": jsonvariables,
    }

    guid_id = create_dashboard_api_query(create_dashboard_query, customer_name)
    return guid_id

def scorecard_automation():
    try:
        while True:
            new_onboard = get_user_input("Are you onboarding a new ACMS customer? (y/n): ").rstrip().lower()
            if new_onboard == "y" or new_onboard == "n":
                break
            print("Invalid input. Please enter 'y' or 'n'.")
                
        while True:
            ent_or_adv = get_user_input("Does your customer have an Advanced package or an Enterprise package? (a/e): ").rstrip().lower()
            if ent_or_adv == "a" or ent_or_adv == "e":
                break
            else:
                print("Invalid input. Please enter 'a' or 'e'.")
    except KeyboardInterrupt:
        sys.exit()
    
    if new_onboard == "y":
        if ent_or_adv == "e":
            customer_name, tenant_id, customer_excel_filename = enter_customer_details()
            scorecard_template, customer_scorecard = scorecard_file_check(new_onboard, customer_excel_filename)
            guid_id = check_for_existing_dashboard(customer_name, new_onboard)
            guid_id = create_new_dashboard(customer_name, tenant_id)
            batch_success_rate_query(tenant_id)
            transactional_success_rate(tenant_id)
            transactional_latency_rate(tenant_id)
            db_space_stats(tenant_id)
            db_bloat_stats(tenant_id)
            sftp_storage_space(tenant_id)
            active_profiles(tenant_id)
            peak_volume_batch_throughout_per_hour(tenant_id)
            peak_volume_transactional_throughput_per_hour(tenant_id)
            customer_names_non_rca_list, customer_names_rca_list = retrieve_jira_customer_list(customer_name)
            select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type="(cpgnprov/cpgnreq) P1/S1")
            guid_id = add_stats_to_scorecard_excel(new_onboard, ent_or_adv, guid_id, scorecard_template, customer_scorecard, customer_name, customer_excel_filename)
        elif ent_or_adv == "a":
            customer_name, tenant_id, customer_excel_filename = enter_customer_details()
            scorecard_template, customer_scorecard = scorecard_file_check(new_onboard, customer_excel_filename)
            guid_id = check_for_existing_dashboard(customer_name, new_onboard)
            guid_id = create_new_dashboard(customer_name, tenant_id)
            batch_success_rate_query(tenant_id)
            transactional_success_rate(tenant_id)
            transactional_latency_rate(tenant_id)
            db_space_stats(tenant_id)
            db_bloat_stats(tenant_id)
            sftp_storage_space(tenant_id)
            active_profiles(tenant_id)
            peak_volume_batch_throughout_per_hour(tenant_id)
            peak_volume_transactional_throughput_per_hour(tenant_id)
            customer_names_non_rca_list, customer_names_rca_list = retrieve_jira_customer_list(customer_name)
            select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type="(cpgnprov/cpgnreq) P1/S1")
            guid_id = add_stats_to_scorecard_excel(new_onboard, ent_or_adv, guid_id, scorecard_template, customer_scorecard, customer_name, customer_excel_filename)
    elif new_onboard == "n":
        if ent_or_adv == "e":
            customer_name, tenant_id, customer_excel_filename = enter_customer_details()
            scorecard_template, customer_scorecard = scorecard_file_check(new_onboard, customer_excel_filename)
            guid_id = check_for_existing_dashboard(customer_name, new_onboard)
            batch_success_rate_query(tenant_id)
            transactional_success_rate(tenant_id)
            transactional_latency_rate(tenant_id)
            db_space_stats(tenant_id)
            db_bloat_stats(tenant_id)
            sftp_storage_space(tenant_id)
            active_profiles(tenant_id)
            peak_volume_batch_throughout_per_hour(tenant_id)
            peak_volume_transactional_throughput_per_hour(tenant_id)
            customer_names_non_rca_list, customer_names_rca_list = retrieve_jira_customer_list(customer_name)
            select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type="(cpgnprov/cpgnreq) P1/S1")
            guid_id = add_stats_to_scorecard_excel(new_onboard, ent_or_adv, guid_id, scorecard_template, customer_scorecard, customer_name, customer_excel_filename)
        elif ent_or_adv == "a":
            customer_name, tenant_id, customer_excel_filename = enter_customer_details()
            scorecard_template, customer_scorecard = scorecard_file_check(new_onboard, customer_excel_filename)
            guid_id = check_for_existing_dashboard(customer_name, new_onboard)
            batch_success_rate_query(tenant_id)
            transactional_success_rate(tenant_id)
            transactional_latency_rate(tenant_id)
            db_space_stats(tenant_id)
            db_bloat_stats(tenant_id)
            sftp_storage_space(tenant_id)
            active_profiles(tenant_id)
            peak_volume_batch_throughout_per_hour(tenant_id)
            peak_volume_transactional_throughput_per_hour(tenant_id)
            customer_names_non_rca_list, customer_names_rca_list = retrieve_jira_customer_list(customer_name)
            select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type="(cpgnprov/cpgnreq) P1/S1")
            guid_id = add_stats_to_scorecard_excel(new_onboard, ent_or_adv, guid_id, scorecard_template, customer_scorecard, customer_name, customer_excel_filename)

    return new_onboard, ent_or_adv

def batch_success_rate_query(tenant_id):
    """Batch Email Success Rate in %"""
    print("PULLING THE SCORECARD METRICS\n")
    query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "FROM BatchDeliveryThroughput SELECT 100 * average(delivery_success_on_first + delivery_success_on_retry) / average(delivery_throughput) AS success_rate WHERE hostname LIKE '{tenant_id}-%-prod%' AND hostname NOT LIKE '%-rt%' where hype_instance_build_version not like '10%' AND delivery_throughput > 0 FACET hostname SINCE 30 day ago", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""

    # Send the API request
    requests_batch_success_rate_query = requests.post(url, headers=headers, data=json.dumps({"query": query}), timeout=timeout)
    # Process the API requests_batch_success_rate_query
    if requests_batch_success_rate_query.status_code == 200:
        json_response = json.dumps(requests_batch_success_rate_query.json())
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["hostname"]))
        if not results:
            print(f"{red}No data found with tenant {blue}{tenant_id}.{endc}")
            sys.exit()
        printed_statement = False
        for result in sorted_results:
            if not printed_statement:
                print(f"{blue}SERVICE QUALITY{endc}\n")
                print(f"{under_line}Batch Success Rate{endc}")
                printed_statement = True
            hostname = result["hostname"]
            short_host = shorten_hostname(hostname)
            success = result["success_rate"]
            if success is not None:
                batch_success_rate = f"{round(success):.0f}%"  # Round success to the nearest integer
                print(short_host, end='')
                print(f": {batch_success_rate}")
                batch_success_rate_results.append(f"{short_host}: {batch_success_rate} | ")
        print("\n")
        return batch_success_rate_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Batch Email Success Rate metric, Nerdgraph query failed with a {requests_batch_success_rate_query.status_code}."
        )
    
def transactional_success_rate(tenant_id):
        """Transaction Email Success Rate in %"""
        query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "SELECT average(success_rate) AS 'success percentage' FROM Postgres WHERE hostname LIKE '{tenant_id}-rt-prod%' AND queryName IN ('TX_MSG_ERR_PEND_VOL_PROCESS_RT', 'TX_MSG_ERR_PEND_VOL_PROCESS_ACS') where hype_instance_build_version not like '10%' where success_rate > 1 FACET hostname, channel SINCE 30 DAYS AGO", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""
        requests_transactional_success_rate_query = requests.post(url, headers=headers, data=json.dumps({"query": query}), timeout=timeout)
        # Process the API requests_transactional_success_rate_query
        if requests_transactional_success_rate_query.status_code == 200:
            data = json.loads(requests_transactional_success_rate_query.content)
            results = data["data"]["actor"]["account"]["nrql"]["results"]
            sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"][0]))
            printed_statement = False
            if sorted_results:
                for result in sorted_results:
                    facet = str(result["facet"][0])
                    if not printed_statement:
                        print(f"{under_line}Transactional Success Rate{endc}")
                        printed_statement = True
                    facet = result["facet"]
                    hostname, channel = facet[0], facet[1]
                    short_host = shorten_hostname(hostname)
                    transactional_success = result["success percentage"] / 100
                    transactional_success_formatted = f"{round(transactional_success * 100):.0f}%"  # Round success to the nearest integer
                    hostname_channel_transactional = f"{short_host}-{channel}: {transactional_success_formatted}"
                    print(f"{hostname_channel_transactional}")
                    output_with_host = f"{short_host}-{channel}: {transactional_success_formatted} | "
                    transactional_success_rate_results_per_channel.append(f"{output_with_host}")
            else:
                print(f"{under_line}Transactional Success Rate{endc}")
                print("No RT data found")
                transactional_success_rate_results_per_channel.append("No RT data found.")
            print("\n")
            return transactional_success_rate_results_per_channel
        else:
            # raise an error with a HTTP response code
            raise Exception(
                f"Unable to retrieve the Transaction Email Success Rate metric, Nerdgraph query failed with a {requests_transactional_success_rate_query.status_code}."
            )

def transactional_latency_rate(tenant_id):
    """Transaction Email Latency Rate in secs"""
    query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "SELECT sum(total_events), sum(under_30s) as under_30s, sum(under_30s) / sum(total_events) as perc_under_30s, sum(under_60s) as under_60s, sum(under_60s) / sum(total_events) as perc_under_60s, sum(under_120s) as under_120s, sum(under_120s) / sum(total_events) as perc_under_120s, sum(under_300s) as under_300s, sum(under_300s) / sum(total_events) as perc_under_300s from TransactionMessagingSample where hostname like '{tenant_id}-rt%-prod%' AND hostname not like '%-mkt%' and hostname not like '%-mid%' where hype_instance_build_version NOT LIKE '10%' where under_30s > 1 facet hostname since 30 day ago", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""
    requests_transactional_latency_rate_query = requests.post(url, headers=headers, data=json.dumps({"query": query}),timeout=timeout)
    # Process the API requests_transactional_latency_rate_query
    if requests_transactional_latency_rate_query.status_code == 200:
        data = json.loads(requests_transactional_latency_rate_query.content)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"]))
        printed_statement = False
        if sorted_results:
            for result in sorted_results:
                facet = result.get("facet")
                if facet is None:
                    print("No RT data found.")
                    transactional_latency_results.append("No RT data found.")
                    continue  # continue with next iteration as there's no facet in this result

                hostname = facet
                short_host = shorten_hostname(hostname)

                under_30 = result.get("perc_under_30s")
                sum_total = result.get("sum.total_events")

                # If there's no data or data is 0.00 for a facet
                if under_30 in [None, 0.00] and sum_total in [None, 0.00]:
                    print("No RT data found.")
                    transactional_latency_results.append("No RT data found.")
                    continue  # continue with next iteration as there's no data in this result

                under_30 = under_30 * 100 if under_30 else None

                if not printed_statement:
                    print(f"{under_line}Transactional Latency{endc}")
                    printed_statement = True

                if under_30:
                    transactional_latency_results.append(f"{short_host}: ")
                    under_30_formatted = f"{round(under_30):.0f}%"  # Round percentage to the nearest integer

                    if sum_total == 0.000:
                        transactional_latency_results.remove(f"{short_host}: ")
                        continue

                    sum_total_formatted = f"Total: {round_up_nr_values_to_nearest_integer(sum_total)}/"
                    print(f"{short_host}: {sum_total_formatted}", end='')
                    transactional_latency_results.append(sum_total_formatted)

                    if under_30:
                        print(f"Under 30s: {under_30_formatted}")
                        transactional_latency_results.append(f"Under 30s: {under_30_formatted} | ")

        else:  # if sorted_results is empty
            print(f"{under_line}Transactional Latency{endc}")
            print("No RT data found.")
            transactional_latency_results.append("No RT data found.")
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Transactional Latency Rate metric, Nerdgraph query failed with a {requests_transactional_latency_rate_query.status_code}."
        )

    print("\n")

def db_space_stats(tenant_id):
    """Database Storage Stats"""
    db_space_query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "SELECT latest(provider.freeStorageSpaceBytes.Sum)/1024/1024/1024 AS 'Free storage space GB', (latest(provider.allocatedStorageBytes) - latest(provider.freeStorageSpaceBytes.Sum))/1024/1024/1024 as 'Used Storage Space GB', latest(provider.allocatedStorageBytes)/1024/1024/1024 as 'Total Storage Space GB', 100 * (latest(provider.allocatedStorageBytes) - latest(provider.freeStorageSpaceBytes.Sum)) / latest(provider.allocatedStorageBytes) as '% Used' FROM DatastoreSample WHERE displayName LIKE '{tenant_id}-%-prod%' AND displayName RLIKE '.*prod([1-9]|[1-9][0-9]|100)$' where hype_instance_build_version not like '10%' AND provider = 'RdsDbInstance' AND providerAccountId = '5856' SINCE 15 minutes ago FACET displayName LIMIT MAX", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""

    # Send the API request
    requests_db_space_stats_query = requests.post(url, headers=headers, data=json.dumps({"query": db_space_query}),timeout=timeout)
    # Process the API requests_db_space_stats_query
    if requests_db_space_stats_query.status_code == 200:
        json_response = json.dumps(requests_db_space_stats_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"]))
        
        printed_statement = False

        for each_result in sorted_results:
            if not printed_statement:
                print(f"{blue}CAPACITY & USAGE{endc}\n")
                print(f"{under_line}Database Storage{endc} - {yellow}If a database instance is missing from storage metrics, it may have an unconventional name, such as from a snapshot or restore. Check the customer's dashboard for unlisted instances.{endc}")
                printed_statement = True
            
            hostname = each_result["facet"]
            short_host = shorten_hostname(hostname)
            print(f"{short_host}: ",end='')
            
            db_storage_space_results.append(f"{short_host}: ")
            used_space_perc = each_result["% Used"]
            used_space = each_result["Used Storage Space GB"]
            total_space = each_result["Total Storage Space GB"]
            disk_space_perc_used = ""

            total_space_formatted = format_storage_values(total_space)
            used_space_formatted = format_storage_values(used_space)
                
            if used_space_perc:
                disk_space_perc_used = f"{used_space_perc:.0f}%"
            print_line = f" {total_space_formatted}/{used_space_formatted}/{disk_space_perc_used} |"
            print(print_line.strip('| '))
            db_storage_space_results.append(print_line.strip(' '))
            
        print("\n")
        return db_storage_space_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Database Storage metric, Nerdgraph query failed with a {requests_db_space_stats_query.status_code}."
        ) 

def db_bloat_stats(tenant_id):
    """Database Bloat Stats"""
    db_bloat_query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "SELECT latest(table_bloat_name), latest(table_wasted_gb) as wasted_GB FROM Postgres WHERE queryName = 'PG_TABLE_BLOAT' and linuxDistribution LIKE '%Debian%' AND hostname LIKE '{tenant_id}-%-prod%' where hype_instance_build_version not like '10%' facet hostname order by `table_wasted_gb` desc since 15 minutes ago", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""
    # Send the API request
    requets_db_bloat_query = requests.post(url, headers=headers, data=json.dumps({"query": db_bloat_query}),timeout=timeout)
    # Process the API requets_db_bloat_query
    if requets_db_bloat_query.status_code == 200:
        json_response = json.dumps(requets_db_bloat_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"]))
        printed_statement = False
        if not sorted_results:
            if not printed_statement:
                print(f"{under_line}Database Bloat{endc}{yellow} - Please note that reporting of bloat occurs only when the bloat size equals 20% of an individual table's size, rather than the entire database.{endc}")
                print("No bloat found in the database.")
                print("\n")
                printed_statement = True
                db_bloat_results.append("No bloat found in the database.")
        for each_result in sorted_results:
            if not printed_statement:
                print(f"{under_line}Database Bloat{endc}")
                printed_statement = True
            hostname = each_result["facet"]
            bloat_table_name = each_result["latest.table_bloat_name"]
            wasted_gigabytes = each_result["wasted_GB"]
            short_host = shorten_hostname(hostname)
            reported_bloat = f"{short_host}: {bloat_table_name} Wasted GB:{wasted_gigabytes:.0f}GB "
            print(reported_bloat)
            db_bloat_results.append(f"{reported_bloat.replace(hostname, '').strip()} |")
        print("\n")
        return db_bloat_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Database Bloat metric, Nerdgraph query failed with a {requets_db_bloat_query.status_code}."
        ) 

def sftp_storage_space(tenant_id):
    """SFTP Storgae Stats"""
    sftp_space_query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "FROM SftpUsageSample SELECT hostname, (allocated / 1024 / 1024 / 1024) as 'Allocated (GB)', (allocated / 1024 / 1024 / 1024) - (current / 1024 / 1024 / 1024) as 'Available (GB)', (current / 1024 / 1024 / 1024) as 'Used (GB)', ((current * 100) / allocated) as '% Used' since 15 minutes ago where hostname like '{tenant_id}-mkt-prod%' where hype_instance_build_version not like '10%'", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}

"""
    # Send the API request
    requests_sftp_storage_space_query = requests.post(url, headers=headers, data=json.dumps({"query": sftp_space_query}),timeout=timeout)
    # Process the API requests_sftp_storage_space_query
    if requests_sftp_storage_space_query.status_code == 200:
        json_response = json.dumps(requests_sftp_storage_space_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["hostname"]))
        printed_statement = False
        for each_result in sorted_results:
            if not printed_statement:
                print(f"{under_line}SFTP Storage{endc}")
                printed_statement = True
            hostname = each_result["hostname"]
            short_host = shorten_hostname(hostname)
            allocated_gb = each_result["Allocated (GB)"]
            used_gb = each_result["Used (GB)"]
            used_gb_perc = each_result["% Used"]
            sftp_results = f"{short_host}: {allocated_gb:.2f}GB/{used_gb:.2f}GB/{used_gb_perc:.2f}% | "
            print(sftp_results.strip(' |'))
            sftp_storage_space_results.append(sftp_results)
        print("\n")
        return sftp_storage_space_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Batch SFTP Storage metric, Nerdgraph query failed with a {requests_sftp_storage_space_query.status_code}."
        ) 

def active_profiles(tenant_id):
    """Active Profile Stats"""
    active_profiles_query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "SELECT latest(current) FROM ActiveProfilesSample where hostname like '{tenant_id}-mkt-prod%' where hype_instance_build_version not like '10%' facet hostname since 30 days ago limit MAX", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""
    # Send the API request
    requests_active_profile_count_query = requests.post(url, headers=headers, data=json.dumps({"query": active_profiles_query}),timeout=timeout)
    # Process the API requests_active_profile_count_query
    if requests_active_profile_count_query.status_code == 200:
        json_response = json.dumps(requests_active_profile_count_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"]))
        printed_statement = False
        for each_result in sorted_results:
            if not printed_statement:
                print(f"{under_line}Active Profiles{endc}")
                printed_statement = True
            hostname = each_result["facet"]
            short_host = shorten_hostname(hostname)
            profile_count = each_result["latest.current"]
            formatted_active_profiles = format_nr_values(profile_count)
            format_nr_values(profile_count)
            print(f"{short_host}: {formatted_active_profiles}")
            active_profiles_results.append(f"{short_host}: {formatted_active_profiles} | ")
        print("\n")
        return active_profiles_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Active Profile Count metric, Nerdgraph query failed with a {requests_active_profile_count_query.status_code}."
        ) 

def peak_volume_batch_throughout_per_hour(tenant_id):
    """Peak Volume Batch Throughput Stats"""
    active_profiles_query = f"""
{{
  actor {{
    account(id: 1209327) {{
      nrql(query: "FROM BatchDeliveryThroughputDetail SELECT sum(total) as 'Peak Throughput' WHERE hostname LIKE '{tenant_id}-%-prod%' AND hostname NOT LIKE '%rt%' and total IS NOT NULL FACET hostname, dateOf(timestamp) as 'Date', hourOf(timestamp) as 'Hour UTC' where hype_instance_build_version NOT LIKE '10%' where total > 1 SINCE 30 days ago limit MAX", timeout: {timeout}) {{
        results
      }}
    }}
  }}
}}
"""
    # Send the API request
    requests_peak_volume_batch_throughput_query = requests.post(url, headers=headers, data=json.dumps({"query": active_profiles_query}),timeout=timeout)
    # Process the API requests_peak_volume_batch_throughput_query
    if requests_peak_volume_batch_throughput_query.status_code == 200:
        json_response = json.dumps(requests_peak_volume_batch_throughput_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"][0]))
        top_peak_batch_result = {}

        for each_result in sorted_results:
            facet_list = each_result["facet"]
            hostname = facet_list[0]
            total_events = each_result["Peak Throughput"]

            if hostname not in top_peak_batch_result or total_events > top_peak_batch_result[hostname]:
                top_peak_batch_result[hostname] = total_events

        if top_peak_batch_result:
            print(f"{under_line}Peak Volume Batch Throughput Per Hour{endc}")

            for hostname, peak_batch in top_peak_batch_result.items():
                short_host = shorten_hostname(hostname)
                formatted_peak_batch = round_up_nr_values_to_nearest_integer(peak_batch)
                batch_throughput = f"{short_host}: {formatted_peak_batch} "
                print(batch_throughput)
                peak_volume_batch_results.append(f"{batch_throughput} | ")

            print("\n")
        return peak_volume_batch_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Peak Volume Batch Throughput Per Hour metric, Nerdgraph query failed with a {requests_peak_volume_batch_throughput_query.status_code}."
        )
print("\n")

def peak_volume_transactional_throughput_per_hour(tenant_id):
    active_profiles_query = f"""
    {{
      actor {{
        account(id: 1209327) {{
          nrql(query: "SELECT sum(total_events) from TransactionMessagingSample where hostname like '{tenant_id}-%-prod%' and hostname not like '%-mkt%' and hostname not like '%-mid%' where hype_instance_build_version NOT LIKE '10%' where total_events > 1 FACET hostname, dateOf(timestamp) as 'Date', hourOf(timestamp) as 'Hour UTC' SINCE 30 days ago limit MAX", timeout: {timeout}) {{
            results
          }}
        }}
      }}
    }}
    """
    # Send the API request
    requests_peak_volume_transactional_throughput_per_hour_query = requests.post(url, headers=headers, data=json.dumps({"query": active_profiles_query}),timeout=240)

    if requests_peak_volume_transactional_throughput_per_hour_query.status_code == 200:
        json_response = json.dumps(requests_peak_volume_transactional_throughput_per_hour_query.json(), indent=10)
        data = json.loads(json_response)
        results = data["data"]["actor"]["account"]["nrql"]["results"]
        sorted_results = sorted(results, key=lambda x: sort_results_by_hostname_number(x["facet"][0]))

        top_peak_trans_result = {}

        for each_result in sorted_results:
            hostname = each_result["facet"][0]
            total_events = each_result["sum.total_events"]

            if hostname not in top_peak_trans_result or total_events > top_peak_trans_result[hostname]:
                top_peak_trans_result[hostname] = total_events

        if top_peak_trans_result:
            print(f"{under_line}Peak Volume Transactional Throughput Per Hour{endc}")

            for hostname, peak_trans in top_peak_trans_result.items():
                short_host = shorten_hostname(hostname)
                formatted_peak_trans = round_up_nr_values_to_nearest_integer(peak_trans)
                transactional_throughput = f"{short_host}: {formatted_peak_trans} "
                print(transactional_throughput)
                peak_volume_transactional_throughput_per_hour_results.append(f"{transactional_throughput} | ")
        else:
            print(f"{under_line}Peak Volume Transactional Throughput Per Hour{endc}")
            print("No RT data found.")
            peak_volume_transactional_throughput_per_hour_results.append("No RT data found.")
            print("\n")
        return peak_volume_transactional_throughput_per_hour_results
    else:
        # raise an error with a HTTP response code
        raise Exception(
            f"Unable to retrieve the Peak Volume Transactional Throughput metric, Nerdgraph query failed with a {requests_peak_volume_transactional_throughput_per_hour_query.status_code}."
        )

def business_days(start_date, end_date):
    return len(pd.bdate_range(start_date, end_date))

def rca_resolved_within_three_days(issues):
    resolved_tickets = [issue for issue in issues if issue["fields"].get("created") and issue["fields"].get("resolutiondate")]

    resolved_created_dates = [datetime.strptime(str(issue["fields"]["created"]), "%Y-%m-%dT%H:%M:%S.%f%z") for issue in resolved_tickets if str(issue["fields"].get("created"))]
    resolved_resolved_dates = [datetime.strptime(str(issue["fields"]["resolutiondate"]), "%Y-%m-%dT%H:%M:%S.%f%z") for issue in resolved_tickets if str(issue["fields"].get("resolutiondate"))]

    resolved_within_three_days_count = 0
    for created_date, resolved_date, issue in zip(resolved_created_dates, resolved_resolved_dates, resolved_tickets):
        if business_days(created_date.date(), resolved_date.date()) <= 3:
            resolved_within_three_days_count += 1
        else:
            print(f"Ticket not resolved within 3 business days: {issue['key']}")

    return resolved_within_three_days_count

def ticket_date_parameters():
    # Get the current date and subtract 1 day
    now = datetime.now() - timedelta(days=1)

    # Calculate the start and end dates for the previous month
    last_month_end = datetime(now.year, now.month, 1) - timedelta(days=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    start_date = last_month_start
    end_date = last_month_end

    # Check for leap year and handle February accordingly
    if calendar.isleap(now.year) and now.month == 3:
        start_date = start_date - timedelta(days=1)

    # Format the date strings in JIRA format (yyyy-mm-dd)
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")

    return start_date_str, end_date_str
    
def retrieve_jira_ticket_count(customer_name, ticket_type = "(cpgnprov/cpgnreq) lower than P1/S1"):
    # Encode the customer name to be used in the URL.
    customer_name_encoded = quote_plus(customer_name)

    # Choose the ticket type. If it's "(cpgnprov/cpgnreq) lower than P1/S1", we consider both 'req' and 'prov'. Otherwise, we consider 'csopm'.
    ticket_types = ['req', 'prov'] if ticket_type == "(cpgnprov/cpgnreq) lower than P1/S1" else ['csopm']

    # Prepare to store the downloaded data.
    df_list = []

    # Open a new web session.
    with requests.Session() as s:
        for ticket in ticket_types:
            # Construct the URL for downloading the data.
            report_url = f'{BASE_URL}/{REPORT_IDS[ticket]}.csv?selected_pages=[Customer Label].[{customer_name_encoded}]'
            # Download the data and process it into a DataFrame.
            df = download_and_process_csv(report_url, s, user_name, password)
            if df.empty is True:
                print(f"Number of tickets logged: {len(df)}")
                print()
                cso_check = retrieve_csopm_ticket_count(customer_name, ticket_type="csopm")
                if cso_check == 0:
                    cso_ticket_count.append(0)
                    rca_res_count.append(0)
                    break
            # Store the DataFrame for later use.
            df_list.append(df)

    # Combine all the DataFrames into one.
    combined_df = pd.concat(df_list)

    # Replace any "NaN" values in the data with zeros, for the specified columns.
    combined_df[['Issues created', 'Issues resolved', 'Average Resolution']] = combined_df[['Issues created', 'Issues resolved', 'Average Resolution']].fillna(0)
    
    if "Unnamed: 1" in combined_df:
        # Filter the data for only the previous month.
        df_filtered = filter_previous_month(combined_df)


        # Depending on the ticket type, calculate and print the 'Average Resolution' from the previous month.
        if ticket_type == "(cpgnprov/cpgnreq) lower than P1/S1":
            cpn_average_resolution = df_filtered['Average Resolution'].sum()
            tickets_logged = df_filtered['Issues created'].sum()
            print(f"Number of tickets logged: {int(tickets_logged)}")
            print(f"Total Average Resolution for (cpgnprov/cpgnreq) lower than P1/S1 Tickets: {cpn_average_resolution:.2f}")
            cpgn_ticket_count.append(int(tickets_logged))
            cpgn_ticket_avg_res.append(f"{cpn_average_resolution:.2f}")
        else:
            # Find the average resolution and Issues created.
            tickets_logged = df_filtered['Issues created'].sum()
            if tickets_logged == 0.00:
                print("No CSO's logged.")
            # If the average resolution is between 0 and 3, print it. Otherwise, print "No CSO's logged".
            print(f"Number of tickets logged: {int(tickets_logged)}")
            print()
            retrieve_csopm_ticket_count(customer_name, ticket_type="csopm")
            # Call a function to determin number of RCA's completed in 3 days or less.
            cso_ticket_count.append(tickets_logged)

        return df_filtered


def retrieve_csopm_ticket_count(customer_name, ticket_type="csopm"):
    if ticket_type:
        jira_api_endpoint, auth = jira_endpoint()

        # Retrieve Start and End date of the previous month for the tickets
        start_date, end_date = ticket_date_parameters()

        rca_tickets_query =  f'project = CSOPM AND "TechOps Product" = "Adobe Campaign" AND createdDate >= {start_date} AND createdDate <= {end_date} AND "Tenant Type" = Single-tenant AND "Customer Name" = "{customer_name}"'

        # Set API parameters
        params = {
            "jql": rca_tickets_query,
            "maxResults": 10000,
            "fields": ["customfield_30000",
                    "created",
                    "resolutiondate", 
                    "key"]
        }

        response = requests.get(jira_api_endpoint, headers=headers, params=params, auth=auth, timeout=timeout)

        # Check if the request was successful and print the ticket count
        if response.ok:
            issues = response.json()["issues"]

            rca_tickets = [issue for issue in issues if str(issue["fields"].get("customfield_30000"))]

            if rca_tickets is not None:
                # Append the three days RCA count
                resolved_within_three_days_count = rca_resolved_within_three_days(rca_tickets)
                print(f"Ticket Type: {ticket_type.upper()}")
                print(f"Total CSOPM tickets resolved within 3 days: {resolved_within_three_days_count}")
                rca_res_count.append(resolved_within_three_days_count)
                print()

        else:
            raise requests.exceptions.RequestException (f"Error retrieving ticket count: {response.status_code} {response.reason}")



def download_and_process_csv(url, session, username, password):
    download = session.get(url, auth=(username, password))
    download.raise_for_status()  # Ensure we get a successful response
    data = download.content.decode('utf-8')
    df = pd.read_csv(StringIO(data))
    
    if "Unnamed: 1" in df.columns:
        # Convert 'Unnamed: 1' to datetime format
        df['Unnamed: 1'] = pd.to_datetime(df['Unnamed: 1'], format='%b %Y')

    return df

def filter_previous_month(df):
    # Get the previous month
    previous_month = datetime.now() - relativedelta(months=1)
    previous_month_start = previous_month.replace(day=1).date()
    previous_month_end = (previous_month_start + relativedelta(months=1, days=-1))

    # Filter rows for the previous month
    mask = (df['Unnamed: 1'].dt.date >= previous_month_start) & (df['Unnamed: 1'].dt.date <= previous_month_end)
    df_filtered = df.loc[mask]
    return df_filtered

def select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type="(cpgnprov/cpgnreq) lower than P1/S1"):
    # Store the input customer name
    user_input = customer_name
    # Choose the appropriate list based on the ticket type
    source_list = customer_names_non_rca_list if ticket_type == "(cpgnprov/cpgnreq) lower than P1/S1" else customer_names_rca_list
    print(f"\nTicket Type: {ticket_type.upper()}")

    # If the input customer name is in the source list
    if user_input in source_list:
        # Call the retrieve_jira_ticket_count function with the input customer name
        retrieve_jira_ticket_count(user_input, ticket_type=ticket_type)

    else:
        # Find the best match and its score using the find_best_match function
        best_match, best_score = find_best_match(user_input, source_list)
        
        # If the best match score is between 90 and 100 (inclusive)
        if 90 <= best_score <= 100:
            # Call the retrieve_jira_ticket_count function with the best match
            retrieve_jira_ticket_count(best_match, ticket_type=ticket_type)

        # If the best match score is less than 90
        elif best_score < 90:
            print("Number of tickets logged: 0")
            print()
            print("Ticket Type: CSOPM")
            print("Total CSOPM tickets resolved within 3 days: 0")
            cso_ticket_count.append(0)
            rca_res_count.append(0)

        # If there is a best match
        elif best_match:
            # Match the casing of the user input to the best match using the match_casing function
            matched_name = match_casing(user_input, best_match)
            # Call the retrieve_jira_ticket_count function with the matched name
            retrieve_jira_ticket_count(matched_name, ticket_type=ticket_type)

        # If there's no best match found
        else:
            # Call the retrieve_jira_ticket_count function with the best match (None)
            retrieve_jira_ticket_count(best_match, ticket_type=ticket_type)

        # If the ticket type is (cpgnprov/cpgnreq) lower than P1/S1 and the best match is not in the customer_names_rca_list
        if ticket_type == "(cpgnprov/cpgnreq) lower than P1/S1":
            if best_match not in customer_names_rca_list:
                pass
        
        # If the ticket type is not (cpgnprov/cpgnreq) lower than P1/S1 and the best match is not in the customer_names_non_rca_list
        else:
            if best_match not in customer_names_non_rca_list:
                pass

        # Return the best match
        return best_match
    
def retrieve_jira_customer_list(customer_name):
    jira_api_endpoint, auth = jira_endpoint()

    # Retrieve Start and End date of the previous month for the tickets
    start_date, end_date = ticket_date_parameters()

    # Set the JIRA JQL search queries to filter tickets by project, customer name, and date range
    customer_list_query = f'"RN Customer Name/s" is not EMPTY AND project in (CPGNREQ, CPGNPROV) AND createdDate >= {start_date} AND createdDate <= {end_date}'
    rca_completed_query = f'project = CSOPM AND "TechOps Product" = "Adobe Campaign" AND createdDate >= {start_date} AND createdDate <= {end_date} AND "Tenant Type" = Single-tenant AND "Customer Name" is not EMPTY'
    
    max_iterations = 5
    loop_counter = 0
    max_results = 1000
    start_at = 0
    customer_names_non_rca_list = set()
    customer_names_rca_list = set()


    progress_bar_1 = tqdm(total=max_iterations, unit='iterations', desc='Retrieving customer names from JIRA for (cpgnprov/cpgnreq) lower than P1/S1 tickets')
    while loop_counter < max_iterations:
        params = {
            "jql": customer_list_query,
            "maxResults": max_results,
            "startAt": start_at,
            "fields": ["customfield_13312","customfield_21300"]
        }

        response = requests.get(jira_api_endpoint, headers=headers, params=params, auth=auth, timeout=timeout)

        if response.ok:
            response_json = response.json()
            issues = response_json["issues"]

            for issue in issues:
                if "customfield_21300" in issue["fields"]:
                    field_value = issue["fields"].get("customfield_21300")
                    if field_value is not None and len(field_value) > 0:  # Check if the list is not empty
                        customer_names_non_rca_list.add(str(field_value[0]))

            start_at += max_results
        else:
            raise requests.exceptions.RequestException(f"Error retrieving customer names: {response.status_code} {response.reason}")
        
        loop_counter += 1
        progress_bar_1.update(1)

    progress_bar_1.close()

    loop_counter = 0
    start_at = 0
    max_iterations = 5
    max_results = 1000


    progress_bar_2 = tqdm(total=max_iterations, unit='iterations', desc='Retrieving customer names from JIRA for CSOPM tickets')
    while loop_counter < max_iterations:
        params = {
            "jql": rca_completed_query,
            "maxResults": max_results,
            "startAt": start_at,
            "fields": ["customfield_30000"]
        }

        response = requests.get(jira_api_endpoint, headers=headers, params=params, auth=auth, timeout=timeout)

        if response.ok:
            response_json = response.json()
            issues = response_json["issues"]

            for issue in issues:
                if "customfield_30000" in issue["fields"]:
                    if issue["fields"].get("customfield_30000") is not None:
                        customer_names_rca_list.add(issue["fields"]["customfield_30000"].lstrip())

            start_at += max_results
        else:
            raise requests.exceptions.RequestException(f"Error retrieving customer names: {response.status_code} {response.reason}")

        loop_counter += 1
        progress_bar_2.update(1)

    progress_bar_2.close()

    customer_names_non_rca_list = sorted(list(set(customer_names_non_rca_list)))
    customer_names_rca_list = sorted(list(set(customer_names_rca_list)))

    select_customer_name(customer_name, customer_names_non_rca_list, customer_names_rca_list, ticket_type = "(cpgnprov/cpgnreq) lower than P1/S1")
    return customer_names_non_rca_list, customer_names_rca_list



def jira_endpoint():
    # Set the JIRA API endpoint and headers
    jira_api_endpoint = "https://jira.corp.adobe.com/rest/api/2/search"
    auth = requests.auth.HTTPBasicAuth(user_name, password)
    return jira_api_endpoint, auth

def set_alignment_and_width(worksheet):
    for col in worksheet.iter_cols(min_row=1, max_row=worksheet.max_row):
        for cell in col:
            cell.alignment = Alignment(wrapText=True, vertical='top', horizontal='left')

    for col in worksheet.columns:
        column_letter = col[3].column_letter
        max_length = 0
        for cell in col[2:]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width
        worksheet.freeze_panes = "B1"

def copy_formatting(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = copy(src_cell.number_format)
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def write_scorecard_stats(wb, new_onboard, ent_or_adv, guid_id, customer_name, customer_excel_filename):
    workbook = wb
    worksheet = workbook.active
    last_month_col_index = None

    if new_onboard == "y":
        if ent_or_adv == "e":
            if 'Advanced' in workbook:
                del workbook['Advanced']
            worksheet = workbook['Enterprise']
        elif ent_or_adv == "a":
            worksheet = workbook['Advanced']
            del workbook['Enterprise']
            worksheet.merge_cells('A3:C3')
            worksheet['A3'] = capitalise_customer_name(customer_name)

        for row in worksheet.iter_rows():
            for index, cell in enumerate(row):
                if index == 2 and "Current" in str(cell.value):
                    cell.value = datetime.now().strftime("%B %Y")
                    cell.font = Font(name=cell.font.name, size=cell.font.size, color="FFFFFFFF")
    
    elif new_onboard == "n":
        for col_num in range(3, worksheet.max_column + 1):
            col = worksheet[get_column_letter(col_num)]
            if col[3].value is not None and any(month in col[3].value for month in calendar.month_name[1:]):
                last_month_col_index = col_num

        if last_month_col_index is not None:
            next_column_index = last_month_col_index + 1
        else:
            next_column_index = 3

        next_column_letter = get_column_letter(next_column_index)
        worksheet.insert_cols(next_column_index)
        new_column = worksheet[next_column_letter + '4']
        current_month = datetime.now().strftime("%B %Y")
        new_column.value = current_month
    
    worksheet['A3'] = capitalise_customer_name(customer_name)

    cell_values = [
        f"{' '.join(batch_success_rate_results)}",
        f"{' '.join(transactional_success_rate_results_per_channel)}" if transactional_success_rate_results_per_channel else "",
        f"{' '.join(transactional_latency_results)}" if transactional_latency_results else "",
        f"{' '.join(db_storage_space_results)}",
        f"{' '.join(db_bloat_results)}",
        f"{' '.join(sftp_storage_space_results)}",
        f"{' '.join(active_profiles_results)}",
        f"{' '.join(peak_volume_batch_results)}",
        f"{' '.join(peak_volume_transactional_throughput_per_hour_results)}",
        f"{cpgn_ticket_count[0] + cso_ticket_count[0]}" if cpgn_ticket_count or cso_ticket_count else "",
        f"{cpgn_ticket_avg_res[0]}" if cpgn_ticket_avg_res else "",
        f"{cso_ticket_count[0]}" if cso_ticket_count else "",
        f"{rca_res_count[0]}" if rca_res_count else "",
    ]

    rows_to_write = [5, 6, 7, 9, 10, 11, 12, 13, 14, 19, 20, 21, 22]

    for i, value in enumerate(cell_values):
            row_to_write = rows_to_write[i]
            if new_onboard == "y":
                worksheet['C' + str(row_to_write)] = value
            elif new_onboard == "n":
                worksheet[next_column_letter + str(row_to_write)] = value

            if new_onboard == "n":
                for row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
                    last_month_cell = row[last_month_col_index - 1]
                    new_cell = row[next_column_index - 1]
                    copy_formatting(last_month_cell, new_cell)

                for column in [worksheet.max_column, worksheet.max_column - 1]:
                    src_cell = worksheet.cell(row=2, column=2)
                    dst_cell = worksheet.cell(row=1, column=column)
                    copy_formatting(src_cell, dst_cell)

                for row in worksheet.iter_rows(min_row=1, max_row=3):
                    for cell in row:
                        src_cell = worksheet.cell(row=2, column=2)
                        dst_cell = cell
                        copy_formatting(src_cell, dst_cell)
            set_alignment_and_width(worksheet)

    if guid_id is not None:
        print(f"Click the link below to access the dashboard for {customer_name}.\n{dash_url(guid_id, customer_name)}")
    workbook.save(f'{customer_excel_filename}_scorecard.xlsx')

def scorecard_file_check(new_onboard, customer_excel_filename):
    """Checks if the scorecard template or customer file exists"""
    scorecard_template = 'TEMPLATE - Service Review Scorecard.xlsx'
    customer_scorecard = f'{customer_excel_filename}_scorecard.xlsx'

    if new_onboard == "y":
        print(f"Checking for {scorecard_template}")
        if os.path.isfile(scorecard_template):
            print(f"{green}{scorecard_template} exists!{endc}")
            return scorecard_template, scorecard_template
        raise FileNotFoundError(f"{red}No scorecard template found with the name `{scorecard_template}`.{endc}")
    
    if new_onboard == "n":
        if os.path.isfile(customer_scorecard):
            print(f"{green}{customer_scorecard} exists!{endc}")
            return customer_scorecard, customer_scorecard
        raise FileNotFoundError(f"{red}The file '{customer_excel_filename}_scorecard.xlsx' was not found.{endc}")
    

def add_stats_to_scorecard_excel(new_onboard, ent_or_adv, guid_id, scorecard_template, customer_scorecard, customer_name, customer_excel_filename):
    """Write Scorecard Stats to Excel"""
    if new_onboard == "y":
            wb = openpyxl.load_workbook(scorecard_template)
            print(f"\n{green}creating {customer_excel_filename}_scorecard.xlsx {endc}\n")
            write_scorecard_stats(wb, new_onboard, ent_or_adv, guid_id, customer_name, customer_excel_filename)
    if new_onboard == "n":
        wb = openpyxl.load_workbook(customer_scorecard)
        print(f"\n{green}writing scorecard metrics to {customer_scorecard}{endc}\n")
        write_scorecard_stats(wb, new_onboard, ent_or_adv, guid_id, customer_name, customer_excel_filename)

def main():
        scorecard_automation()
if __name__ == "__main__":
    main()